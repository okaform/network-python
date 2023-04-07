'''This script will generate a simple report of pings.
 RIght now it is onlye 1 ping tries and then it moves on.
It will also add it to the excel file.'''

import sys, re, os
import xlsxwriter
import getOutput
import openpyxl
from datetime import date, datetime  
import subprocess
'''I am IMPORTING LIKE This because of the nature of the VDI not ening persistent. 
I am adding this directory to the python path so that python can find the library.'''
sys.path.append('N:/Python Libraries')

from tabulate import tabulate
#from openpyxl import Workbook


start_time = datetime.now()
as_id = "as_cz507f" #input("Enter your as_id")
as_pass = "vP76loNxtDjqpnzS" #input("Enter your as_password")


#workbook = xlsxwriter.Workbook( "report/report " + str(datetime.now().strftime("%b-%d-%y [%H-%M]")) +".xlsx")
#sheet1 = workbook.add_worksheet("REPORT")
#bold = workbook.add_format({'bold': True})
row = 1 #row and columns must be at least one with openpyxl
col = 1
#sheet1.write(row, col, "Date", bold)
#sheet1.write(row, col+1, "Hostname", bold)
#sheet1.write(row, col+2, "IP Address", bold)
#sheet1.write(row, col+3, "Switch Model", bold)
#sheet1.write(row, col+4, "Installed Version", bold)

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title="REPORT"
file_name = "N:\\Report\\ping\\report " + str(datetime.now().strftime("%b-%d-%y [%H-%M]")) +".xlsx"
#workbook.save(file_name)
sheet1.cell(row=row, column=col, value="Date")
sheet1.cell(row=row, column=col+1, value="Hostname")
sheet1.cell(row=row, column=col+2, value="IP Address")
sheet1.cell(row=row, column=col+3, value="Status")





print("\n\nPaste the IP Addresses you to do a ping sweep for. \n"+
"When you are done. Hit ENTER, Ctrl + z, and ENTER in that order to end:\n")
ip_Address = sys.stdin.readlines() #This will read multiple lines


data = []

for i in range(len(ip_Address)):
      '''STEP ONE: Get Date '''
      sheet1.cell(row + i + 1, col, datetime.now().strftime("%b-%d-%y %H:%M:%S"))
      workbook.save(file_name) #save after every function call
      '''STEP TWO: Get Hostname '''
      sheet1.cell(row + i + 1, col+1, getOutput.hostname(ip_Address[i]))
      workbook.save(file_name) #save after every function call
        
      '''STEP THREE: Get ip Address '''
      sheet1.cell(row + i + 1, col+2, ip_Address[i])
      workbook.save(file_name) #save after every function call      

      '''STEP FOUR: DO THE PING '''
      status = getOutput.will_ping(ip_Address[i])
      sheet1.cell(row + i + 1, col+3, status)
      workbook.save(file_name) #save after every function call   
      data.append([ip_Address[i], status])


      if i != len(ip_Address) - 1: #do not print this on the last switch
          print("\n-----------------------------------------------------------")    
          print("  ------------- Moving on to " +str(ip_Address[i + 1]).strip() + " ----------------")
          print("-----------------------------------------------------------\n\n")         

        
print(tabulate(data, headers=["IP Address", "Status"]))   
workbook.save(file_name)
#End of Script
elapsed_time = datetime.now() - start_time #to calculate the total elapsed time the script took to run
print("This script took approximately {}".format(elapsed_time))



'''Write fix for if they post with hostnames. It will be as using regex to see if the format matches ip address format. IF it doesn't then use hostname generator to get the ip address
SUPER FUN STUFF'''