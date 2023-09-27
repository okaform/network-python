import sys, re, os
import openpyxl
import nslookup
from openpyxl.styles import Alignment, PatternFill, Font
from vlan_dict import vlan_id_to_no
from cab import cab_dict
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import SSHException
from netmiko.ssh_exception import AuthenticationException
from datetime import date, datetime  



#print("This is sys.argv 0 ->",sys.argv[1])
#date.today())+
#str(datetime.now().strftime("%m-%d-%y-[%H-%M]"))
start_time = datetime.now()

reg = re.compile(r"[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4}")#regEX for Mac address
regIP = re.compile(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}")#find regEx for IPs 
regVlan = re.compile(r"\b\d{1,4}\b") 
''' -------------------------------------
    --------- CREDENTIAL OPTIONS --------
    ------------------------------------- '''
as_id = "as_cz507f" #input("Insert your as_id: ")
as_pass = "DBf058sNjmqrCRaK" #input("\nInsert your as_password: ")
en_secret = "DBf058sNjmqrCRaK" #input("\nInsert your secret password:")
#This is useful for switches without management template. It will use the enable secret
if en_secret == '':
    en_secret = as_pass

#print(en_secret)
''' -----------------------------------
    ------- switch IP address -----
    ----------------------------------- '''
#switch to get the mac address from
mac_switch = "10.51.0.19" #input("Type IP address of switch to get mac address: ")
ip_switch = "10.51.5.202" #input("Type IP address of switch to get IP address: ")
''' ---------------------------------------
    ---------- PASTE PORT NUMBERS ---------
    --------------------------------------- '''
print("\n\nPaste port numbers. \n"+
"When you are done. Hit ENTER, Ctrl + z, and ENTER in that order to end:\n")

port_no = sys.stdin.readlines() #This will read multiple lines

''' ------------------------------------------
    ---------- CONNECT to MAC_ SWITCH ---------
    ------------------------------------------ '''
net_dev_mac = {"host":mac_switch,
                    "username":as_id,
                    "password":as_pass, 
                    "device_type":"cisco_ios",
                    "secret":en_secret
                    }
try:
    con_mac = ConnectHandler(**net_dev_mac)
except:
    issue_st = "\nThere is a problem with your log in credentials for " +str(mac_switch).strip()+". Please check your username or password.\n"
    print(issue_st)
    logs_err = open("ZZ-error-logs.txt", mode="a")
    logs_err.write(issue_st)
    exit() #end the function

            
''' ------------------------------------------
    ---------- CONNECT TO IP SWITCH ---------
    ------------------------------------------ '''    
net_dev_ip = {"host":ip_switch,
                    "username":as_id,
                    "password":as_pass, 
                    "device_type":"cisco_ios",
                    "secret":en_secret
                    }
try:
    con_ip = ConnectHandler(**net_dev_ip)
except:
    issue_st = "\nThere is a problem with your log in credentials for " +str(ip_switch).strip()+". Please check your username or password.\n"
    print(issue_st)
    logs_err = open("ZZ-error-logs.txt", mode="a")
    logs_err.write(issue_st)
    exit() #end the function


''' ------------------------------------------
    ---------- SPREADSHEET IMPLEMENTATION ---------
    ------------------------------------------ '''
bold_font = Font(bold=True) #To make the font bold
# Create an Alignment object for top alignment and wrap text
top_alignment_wrap = Alignment(vertical='top', wrap_text=True)

row = 1 #row and columns must be at least one with openpyxl
col = 1
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title="REPORT"
file_name = "N:\Report\DC_server_info\ " + str(datetime.now().strftime("%b-%d-%y [%H-%M]")) +".xlsx"
#workbook.save(file_name)
sheet1.cell(row=row, column=col, value="Device IP").font = bold_font
sheet1.cell(row=row, column=col+1, value="Port Name").font = bold_font
sheet1.cell(row=row, column=col+2, value="Description").font = bold_font
sheet1.cell(row=row, column=col+3, value="IP/MAC").font = bold_font
sheet1.cell(row=row, column=col+4, value="FQDN").font = bold_font
sheet1.cell(row=row, column=col+5, value="CAB").font = bold_font
sheet1.cell(row=row, column=col+6, value="Default Incident Group").font = bold_font
sheet1.cell(row=row, column=col+7, value="Assignment Group").font = bold_font
sheet1.cell(row=row, column=col+8, value="Task Implementation Services").font = bold_font
sheet1.cell(row=row, column=col+9, value="Blackout Monitor").font = bold_font
sheet1.cell(row=row, column=col+10, value="Post Deployment UAT").font = bold_font
sheet1.cell(row=row, column=col+11, value="VLAN Name").font = bold_font
sheet1.cell(row=row, column=col+12, value="VLANID").font = bold_font

''' ------------------------------------------
    ---------- ACTUAL IMPLEMENTATION ---------
    ------------------------------------------ '''   
        
for i in range(len(port_no)):    
    '''PORT NAME'''
    sheet1.cell(row+i+1, col+1, str(port_no[i]).strip()).alignment = top_alignment_wrap
    if len(port_no) != 1: #this takes care of empty lines so they are not looked at
        '''VLAN NAME and VLAN ID'''
        command_to_run = "show run int " +str(port_no[i]).strip() + " | inc switchport"
        vlan_found = con_mac.send_command(command_to_run)
        vlan_found_reg = re.findall(regVlan, vlan_found)
        vlan_id = ''#this is for the vlan id 
        if len(vlan_found_reg) == 0: #This will help with empty list
            sheet1.cell(row+i+1, col+12, "NO VLAN ID FOUND").alignment = top_alignment_wrap #update the vlan id
        else:
            vlan_id = vlan_found_reg[0].strip()
            sheet1.cell(row+i+1, col+12, str(vlan_id)).alignment = top_alignment_wrap #update the vlan id
            
        if vlan_id in vlan_id_to_no:#find the vlan name from the dictionary
            sheet1.cell(row+i+1, col+11, str(vlan_id_to_no[vlan_id]).strip())#update the vlan name
        else:
            sheet1.cell(row+i+1, col+11, "NO VLAN NAME FOUND").alignment = top_alignment_wrap
            
        '''MAC ADDRESS'''    
        mac_address_found = con_mac.send_command("show mac address-table | inc " +str(port_no[i]))
        mac_address_reg = re.findall(reg, mac_address_found)   
        if len(mac_address_reg) == 0:
            sheet1.cell(row+i+1, col+3, "No Mac Address").alignment = top_alignment_wrap
            print("No Mac Address")
        
        for mac in mac_address_reg: #take that mac and get an ip.
            ip_found = con_ip.send_command("show ip arp " +str(mac))
            ip_reg = re.findall(regIP, ip_found)
            '''IP Address'''            
            if len(ip_reg) == 0:
                #if no ip is found, then use the mac address
                print("No IP Address: Mac Address:" +str(mac))
                sheet1.cell(row+i+1, col+3, str(mac).strip()) #.alignment = top_alignment_wrap
                sheet1.cell(row+i+1, col+4, "No IP Address") #.alignment = top_alignment_wrap
            else:
                ip_reg_string = '\n'.join(ip_reg)
                sheet1.cell(row+i+1, col+3, ip_reg_string).alignment = top_alignment_wrap
                print(ip_reg)
                ns = []
                for ip in ip_reg:
                    hostname_found = nslookup.nslookup(ip)[1].strip()
                    ns.append(hostname_found) #add new nslookup to the list
                
                '''CAB IMPLEMENTATION''' #take the hostname and look for implementation
                for key in cab_dict:#loop through the dictionary
                    if key in ns[0]: #if key words are found in hostname, return the matching list 
                        print(cab_dict[key])
                        '''CAB'''
                        sheet1.cell(row+i+1, col+5, str(cab_dict[key][0]))
                        '''ASSIGNMENT GROUP'''
                        sheet1.cell(row+i+1, col+7, str(cab_dict[key][1]))
                        '''IMPLEMENTATION SERVICES'''
                        sheet1.cell(row+i+1, col+8, str(cab_dict[key][2]))
                        break
                    
                        
                ns_look = '\n'.join(ns) #split the list to a string
                '''FQDN'''
                sheet1.cell(row+i+1, col+4, ns_look).alignment = top_alignment_wrap#add to the next column
                    
        
    workbook.save(file_name)    
        
    if i != len(port_no) - 1: #do not print this on the last switch
        print("\n--------------------------------------------------------------------------")    
        print("----------------------- Moving on to " +str(port_no[i + 1]).strip() + " -------------------------")
        print("--------------------------------------------------------------------------\n\n") 
        
print("Successfully Completed. Please check the " + str(file_name) +"in the DC_server_info Folder")

elapsed_time = datetime.now() - start_time #to calculate the total elapsed time the script took to run
print("This script took approximately {}".format(elapsed_time))
        
        