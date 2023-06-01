'''This script will generate a simple report of cisco switches with their current running ver, their installed committed and their model.
It will also add it to the excel file.'''

import sys, re, os
import xlsxwriter
import getOutput
import openpyxl
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import SSHException
from netmiko.ssh_exception import AuthenticationException
from datetime import date, datetime  
#from openpyxl import Workbook


start_time = datetime.now()
as_id = "as_cz507f" #input("Enter your as_id")
as_pass = "4D1nGmZNoKTBJC39" #input("Enter your as_password")


#workbook = xlsxwriter.Workbook( "report/report " + str(datetime.now().strftime("%b-%d-%y [%H-%M]")) +".xlsx")
#sheet1 = workbook.add_worksheet("REPORT")
#bold = workbook.add_format({'bold': True})
row = 1 #row and columns must be at least one with openpyxl
col = 1
#sheet1.write(row, col, "Date", bold)
#sheet1.write(row, col+1, "Hostname", bold)
#sheet1.write(row, col+2, "IP Address", bold)
#sheet1.write(row, col+3, "Switch Model", bold)
#sheet1.write(row, col+4, "Installed Version", bold)

workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title="REPORT"
file_name = "N:\Report\cisco-version\ " + str(datetime.now().strftime("%b-%d-%y [%H-%M]")) +".xlsx"
#workbook.save(file_name)
sheet1.cell(row=row, column=col, value="Date")
sheet1.cell(row=row, column=col+1, value="Hostname")
sheet1.cell(row=row, column=col+2, value="IP Address")
sheet1.cell(row=row, column=col+3, value="Switch Model")
sheet1.cell(row=row, column=col+4, value="Installed Version")




print("\n\nPaste the IP Addresses you want prechecks generated. \n"+
"When you are done. Hit ENTER, Ctrl + z, and ENTER in that order to end:\n")
ip_Address = sys.stdin.readlines() #This will read multiple lines



for i in range(len(ip_Address)):
    if len(ip_Address[i]) != 1: #this takes care of empty lines so they are not looked at
        net_dev = {"host":ip_Address[i],
                    "username":as_id,
                    "password": as_pass, 
                    "device_type":"cisco_ios",
                    }
        
        try:
            con = ConnectHandler(**net_dev)
        except:
            print("There is a problem with your log in credentials for " +str(ip_Address[i])+". Please check your username or password")
            sheet1.cell(row + i + 1, col, datetime.now().strftime("%b-%d-%y %H:%M:%S"))
            sheet1.cell(row + i + 1, col+2, ip_Address[i])
            sheet1.cell(row + i + 1, col+3, "Could Not Log In")
            continue #This should break out of the loop
            
        try:    
            con.enable()
        except:
            print("There is a Enable Password issue: " +str(ip_Address[i]).strip()+". Please check the device")
            sheet1.cell(row + i + 1, col, datetime.now().strftime("%b-%d-%y %H:%M:%S"))
            sheet1.cell(row + i + 1, col+2, ip_Address[i])
            sheet1.cell(row + i + 1, col+3, "Could Not Enter Enable Mode ")
            continue #This should break out of the loop
        '''STEP ONE: Get Date '''
        sheet1.cell(row + i + 1, col, datetime.now().strftime("%b-%d-%y %H:%M:%S"))
        workbook.save(file_name) #save after every function call
        
        '''STEP TWO: Get Hostname '''
        sheet1.cell(row + i + 1, col+1, getOutput.hostname(con, ip_Address[i]))
        workbook.save(file_name) #save after every function call
        
        '''STEP THREE: Get ip Address '''
        sheet1.cell(row + i + 1, col+2, ip_Address[i])
        workbook.save(file_name) #save after every function call
        
        '''STEP FOUR: Get Switch Model '''
        sheet1.cell(row + i + 1, col+3, getOutput.switch_model(con))
        workbook.save(file_name) #save after every function call
        
        '''STEP FIVE: Get Install Commited '''
        sheet1.cell(row + i + 1, col+4, getOutput.show_install_committed(con))
        workbook.save(file_name) #save after every function call
        
        
        #getOutput.getSwitchModel(con)
    if i != len(ip_Address) - 1: #do not print this on the last switch
        print("\n-----------------------------------------------------------")    
        print("  ------------- Moving on to " +str(ip_Address[i + 1]).strip() + " ----------------")
        print("-----------------------------------------------------------\n\n")         
        

workbook.save(file_name)
#End of Script
elapsed_time = datetime.now() - start_time #to calculate the total elapsed time the script took to run
print("This script took approximately {}".format(elapsed_time))



'''Write fix for if they post with hostnames. It will be as using regex to see if the format matches ip address format. IF it doesn't then use hostname generator to get the ip address
SUPER FUN STUFF'''