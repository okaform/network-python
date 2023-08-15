import sys, os
import openpyxl
import csv
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import SSHException
from netmiko.ssh_exception import AuthenticationException
from datetime import date, datetime  
from header_to_function import header_to_function_dict
sys.path.append('N:/Python Libraries')
import pandas as pd

''' -------------------------------------------------
    ---------- CONVERT FROM CSV TO XLSX -------------
    -------------------------------------------------- '''
csv_filename = "N:\\Report\\vEdge-to-cEdge\\MFG-NA-8300-R1-M1-I2-I3-I1-LTE-T4-v5.csv"
r1_excel = "N:\\Report\\vEdge-to-cEdge\\r1-temp.xlsx"
data = pd.read_csv(csv_filename) #Load CSV file using pandas
data.to_excel(r1_excel, index=False)


''' ---------------------------------------
    ---------- OPEN EXCEL FILE -------------
    --------------------------------------- '''
try:
    workbook = openpyxl.load_workbook("N:\\Report\\vEdge-to-cEdge\\r1-temp.xlsx") #Open the newly converted excel file
    sheet1 = workbook.active
    col = 1
  
    #print(sheet1.max_column) #to get the size of the column
    for i in range(sheet1.max_column): #loop through the list of header cells
        header = sheet1.cell(row = 1, column = col + i).value #Get the value of the header
        #print(sheet1.cell(row = 1, column = col + i).value)
        if header in header_to_function_dict: #if header value is in the dictionary
            if type(header_to_function_dict[header]) is str: #check if dictionary value is function or string
                value_to_set = header_to_function_dict[header]
                print(value_to_set)
                sheet1.cell(row = 2, column = col + i, value=value_to_set) #Update the coresponding column 
            else:
                value_to_set = header_to_function_dict[header]()
                sheet1.cell(row = 2, column = col + i, value=value_to_set) #Update the coresponding column 
    
    workbook.save("N:\\Report\\vEdge-to-cEdge\\r1-temp.xlsx")
    
except FileNotFoundError:
    print("File not found:", r1_template)
except openpyxl.utils.exceptions.InvalidFileException:
    print("Invalid Excel file format:", r1_template)
    